from django.shortcuts import render
from django.db.models import Count
from django.views.generic import ListView, DetailView, UpdateView
from django.views.generic.edit import CreateView
from .models import *
from datetime import datetime
from datetime import timedelta
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl import load_workbook


def main(request):
    student = Students.objects.all().count()
    period = Period.objects.all()
    questions = Period.objects.annotate(number_of_period=Count('period'))
    context = {
        'title': 'Главная страница',
        'student': student,
        'period': period,
        'questions': questions,
    }
    return render(request, 'main/index.html', context=context)


def view_period(request, period_id):
    period_item = Period.objects.get(pk=period_id)
    students_item = Students.objects.all().filter(keyPeriod=period_id)
    context = {
        'title': period_item.title,
        'period_item': period_item,
        'students_item': students_item,
    }
    return render(request, 'main/period/detail.html', context=context)


class StudentListView(ListView):
    model = Students
    template_name = 'main/index.html'


class StudentDetailView(DetailView):
    model = Students
    template_name = 'main/detail.html'


class StudentCreateView(CreateView):
    model = Students
    template_name = 'main/create.html'
    fields = ['keyPeriod', 'surname', 'name', 'patronymic', 'sex', 'birthday', 'snils']


class StudentUpdateView(UpdateView):
    model = Students
    template_name = 'main/edit.html'
    fields = ['keyPeriod', 'surname', 'name', 'patronymic', 'sex', 'birthday', 'snils']


def export_movies_to_xlsx(request, period_id):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    period_item = Period.objects.get(pk=period_id)
    movie_queryset = Students.objects.all().filter(keyPeriod=period_id)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Movies'

    # Define the titles for columns
    columns = [
        'ID',
        'Title',
        'Description',
        'Length',
        'Rating',
        'Price',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for movie in movie_queryset:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            movie.pk,
            movie.title,
            movie.description,
            movie.length_in_minutes,
            movie.rating,
            movie.price,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def my_print(request):
    from openpyxl import Workbook
    from openpyxl.writer.excel import save_virtual_workbook
    import datetime

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    ws.append([1, 2, 3])
    ws['A2'] = datetime.datetime.now()
    print(save_virtual_workbook(wb))

    print("Content-Type: application/vnd.ms-excel")
    print("Content-Disposition: attachment; filename=test.xlsx")

    return request


def print_me(request, period_id):
    period_item = Period.objects.get(pk=period_id)
    students_item = Students.objects.all().filter(keyPeriod=period_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Movies'
    columns = [
        'ФАМИЛИЯ',
        'ИМЯ',
        'ОТЧЕСТВО',
        'ДАТА РОЖДЕНИЯ',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for movie in students_item:
        row_num += 1
        row = [
            movie.surname,
            movie.name,
            movie.patronymic,
            movie.birthday,
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save("media/list.xlsx")
    context = {
        'title': 'Печать',
        'period_item': period_item,
        'students_item': students_item,
    }
    return render(request, 'main/period/print.html', context=context)

