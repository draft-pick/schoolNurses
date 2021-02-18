from django.shortcuts import render
from django.db.models import Count
from .models import *
from django.views.generic.edit import CreateView
from django.views.generic import ListView, DetailView, UpdateView
from openpyxl import Workbook, load_workbook
from docxtpl import DocxTemplate


def home(request):
    periods = Period.objects.filter(is_active=1)
    students_count = Period.objects.annotate(number_of_period=Count('period'))
    context = {
        'periods': periods,
        'title': 'Главная страница',
        'students_count': students_count,
    }
    return render(request, 'school/index.html', context=context)


class PeriodCreateView(CreateView):
    model = Period
    template_name = 'school/periods/create.html'
    fields = ['title', 'date', 'is_active']


class PeriodDetailView(DetailView):
    model = Period
    template_name = 'school/periods/detail.html'


def view_period(request, period_id):
    period_item = Period.objects.get(pk=period_id)
    students_item = Students.objects.all().filter(keyPeriod=period_id)
    context = {
        'title': period_item.title,
        'period_item': period_item,
        'students_item': students_item,
    }
    return render(request, 'school/periods/open.html', context=context)


class StudentCreateView(CreateView):
    model = Students
    template_name = 'school/students/create.html'
    fields = ['keyPeriod', 'surname', 'name', 'patronymic', 'sex', 'birthday', 'snils']


class StudentDetailView(DetailView):
    model = Students
    template_name = 'school/students/detail.html'


class StudentUpdateView(UpdateView):
    model = Students
    template_name = 'school/students/edit.html'
    fields = ['keyPeriod', 'surname', 'name', 'patronymic', 'sex', 'birthday', 'snils']


def export_xlsx(request, period_id):
    period_item = Period.objects.get(pk=period_id)
    students_item = Students.objects.all().filter(keyPeriod=period_id)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Students'
    columns = [
        'ФАМИЛИЯ',
        'ИМЯ',
        'ОТЧЕСТВО',
        'ДАТА РОЖДЕНИЯ',
        'ПОЛ',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    for students in students_item:
        row_num += 1
        row = [
            students.surname,
            students.name,
            students.patronymic,
            students.birthday,
            students.sex
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save("media/list.xlsx")
    context = {
        'title': 'Печать',
        'period_item': period_item,
        'students_item': students_item,
    }
    return render(request, 'main/period/print.html', context=context)


def export_docx(request, pk):
    student_item = Students.objects.filter(pk=pk)
    for student in student_item:
        doc = DocxTemplate("media/temp_doc.docx")
        context = {
            'surname': student.surname,
            'name': student.name,
            'patronymic': student.patronymic,
        }
        doc.render(context)
        doc.save("media/generated_doc.docx")
    return render(request, 'school/students/print_docx.html')